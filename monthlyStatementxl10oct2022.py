from ctypes import alignment
from datetime import datetime
import datetime
from tkinter import HORIZONTAL, font
from turtle import fillcolor, left
import pandas as pd
from fpdf import FPDF
statementOf=datetime.date.today() + datetime.timedelta(days=-2)
def __init__(self):
    self=self
def makestatement(statementOf):
    print('creating statement')
    import openpyxl
    import os,sys
    from pathlib import Path
    d = datetime.datetime.now()+ datetime.timedelta(days=-2)
    xlsx_file = Path("C:/pdfInvoiceMaker/statement.xlsx")
    wb = openpyxl.load_workbook(xlsx_file,read_only=False,data_only=False)
    from openpyxl.styles.borders import Border, Side
    from openpyxl.styles import Font, Fill,colors,Alignment,PatternFill
    thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
    play_data = wb['statement']

    play_data['E3'].value=d.strftime("%B-%Y")
    d = d.strftime("%B-%Y")    

    with open(f"C:/pdfInvoiceMaker/statement-{d}.txt") as f:
        lines=f.readlines()
        cc=0
        for line in lines:
            if len(line)>5:
                line=(line[:-1])
                c=0
                for datas in line.split(','):
                    
                    if c==0:
                        play_data.cell(row=8+cc, column=1).border = thin_border
                        play_data[f'A{8+cc}'].value=datas
                    elif c==1:
                        play_data.cell(row=8+cc, column=2).border = thin_border
                        play_data[f'B{8+cc}'].value=datas
                    elif c==2:
                        play_data.cell(row=8+cc, column=3).border = thin_border
                        play_data[f'c{8+cc}'].value=float(datas) 
                    elif c==3:
                        play_data.cell(row=8+cc, column=4).border = thin_border
                        play_data[f'd{8+cc}'].value=float(datas) 
                    elif c==4:
                        play_data.cell(row=8+cc, column=5).border = thin_border
                        play_data[f'e{8+cc}'].value=float(datas) 
                    elif c==5:
                        play_data.cell(row=8+cc, column=6).border = thin_border
                        play_data[f'f{8+cc}'].value=float(datas) 
                    elif c==6:
                        play_data.cell(row=8+cc, column=7).border = thin_border
                        play_data.cell(row=8+cc, column=5).border = thin_border
                        play_data[f'g{8+cc}'].value=float(datas)
                        play_data[f'e{8+cc}'].value=float(datas)                  
                    c=c+1
                    wb.save(f'C:/pdfInvoiceMaker/statement-{d}.xlsx')
                    
            cc+=1

    play_data[f'g{9+cc}']=f"=SUM(g8:g{8+cc})"
    play_data[f'f{9+cc}']=f"=SUM(f8:f{8+cc})"
    play_data[f'e{9+cc}']=f"=SUM(e8:e{8+cc})"
    play_data[f'd{9+cc}']=f"=SUM(d8:d{8+cc})"
    play_data[f'c{9+cc}']=f"=SUM(c8:c{8+cc})"
    play_data[f'b{8+cc}']=f'=SMALL(data,ROWS($B$8:B{8+cc}))'

    play_data[f'A{9+cc}']='Total :'
    for i in range(5):
        play_data.merge_cells(f'A{10+cc+i}:b{10+cc+i}')
        play_data.merge_cells(f'd{10+cc+i}:e{10+cc+i}')

    play_data[f'A{10+cc}'].value=f'Total No of invoices:'
    play_data[f'c{10+cc}'].value=f'{cc}'
    #play_data[f'd{10+cc}'].value=f"=SpellNumber(c{10+cc})"

    play_data[f'A{11+cc}'].value=f'Total Amount:'
    play_data[f'c{11+cc}'].value=f'=c{cc+9}'
    #play_data[f'd{11+cc}'].value=f"=SpellNumber(c{11+cc})"

    play_data[f'A{12+cc}'].value=f'Total Tax:'
    play_data[f'c{12+cc}'].value=f'=d{cc+9}'
    #play_data[f'd{12+cc}'].value=f"=SpellNumber(c{12+cc})"

    play_data[f'A{13+cc}'].value=f'Total Debit:'
    play_data[f'c{13+cc}'].value=f'=e{cc+9}'
    #play_data[f'd{13+cc}'].value=f"=SpellNumber(c{13+cc})"

    for i in range(5):
        play_data[f'A{9+cc+i}'].font=Font(size=9,bold=True)
        play_data[f'A{9+cc+i}'].alignment=Alignment(horizontal='left')
        play_data[f'c{9+cc+i}'].font=Font(size=9,bold=True)
        play_data[f'c{9+cc+i}'].alignment=Alignment(horizontal='left')
        play_data[f'd{9+cc+i}'].font=Font(size=9,bold=True)
        play_data[f'd{9+cc+i}'].alignment=Alignment(horizontal='left')


    wb.save(f'C:/pdfInvoiceMaker/statement-{d}.xlsx')
    wb.close()
    print('Uploading and printing statement Done')
    #os.startfile(f'C:/pdfInvoiceMaker/statement-{d}.xlsx','print')
    from pydrive.auth import GoogleAuth
    from pydrive.drive import GoogleDrive
    gauth = GoogleAuth()
    drive = GoogleDrive(gauth)
    f1=drive.CreateFile({'title': f'statement of {d}.xlsx'})
    f1.SetContentFile(f'C:/pdfInvoiceMaker/statement-{d}.xlsx')   
    f1.Upload()
a=makestatement(statementOf)
